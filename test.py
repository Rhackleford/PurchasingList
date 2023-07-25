import pandas as pd
import numpy as np

# Load spreadsheet
xl = pd.read_excel('testjob.xls', engine='xlrd')

# Find the index of the row containing "HARDWARE PARTS"
start_index = xl[xl.apply(lambda row: row.astype(str).str.contains('HARDWARE PARTS').any(), axis=1)].index[0]

# Find the index of the row containing "Packsize Program"
end_index = xl[xl.apply(lambda row: row.astype(str).str.contains('Packsize Program').any(), axis=1)].index[0]

# Remove all rows above the start row and below the end row
xl = xl.loc[start_index:end_index-1]


# Find the index of the row containing "Metal Parts - Cut Length and Qty"
metal_parts_index = xl[xl.apply(lambda row: row.astype(str).str.contains('Metal Parts - Cut Length and Qty').any(), axis=1)].index[0]

# Find the index of the row containing "ACCESSORY PARTS for BUYOUT"
accessory_buyout_index = xl[xl.apply(lambda row: row.astype(str).str.contains('ACCESSORY PARTS for BUYOUT').any(), axis=1)].index[0]

# Remove all rows from "Metal Parts - Cut Length and Qty" to "ACCESSORY PARTS for BUYOUT"
xl = xl.drop(list(range(metal_parts_index, accessory_buyout_index)))
# If column 3 is NaN, empty, or blank, move the value from column 2 to column 3
xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()), xl.columns[3]] = xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()), xl.columns[2]]

# Set the corresponding cell in column 2 to empty
xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()), xl.columns[2]] = ""

# Define a function to check if a value is a number
def is_number(n):
    try:
        float(n)   # Try to convert the string to float
        return True
    except ValueError:   # If ValueError is thrown, it's not a number
        return False

# Check if column 6 contains a numeric value and column 5 is empty or NaN
numeric_mask = xl[xl.columns[6]].apply(is_number) & (xl[xl.columns[5]].isna() | (xl[xl.columns[5]] == '') | (xl[xl.columns[5]].astype(str).str.isspace()))

# Move the value from column 6 to column 5
xl.loc[numeric_mask, xl.columns[5]] = xl.loc[numeric_mask, xl.columns[6]]

# Set the corresponding cell in column 6 to empty
#xl.loc[numeric_mask, xl.columns[6]] = ""


# Define the list of sections
sections = ['HARDWARE PARTS', 'Hinges & Mounting Plates', 'Legrabox & Antaro', 'Metabox, Tandem & Accuride', 'Blum Metal Parts',
            'Closet', 'Other', 'ACCESSORY PARTS for BUYOUT',
            'ADDITIONAL ACCESSORY PARTS', 'Decorative Hardware', 'RECESSED HARDWARE - Install Prior to Shipping',
            'Berenson INTEGRATED PULL PARTS']



# Create 'Section' column and initialize with NaN
xl.insert(0, 'Section', np.nan)

# Iterate over the section names and assign them to 'Section' column
for section in sections:
    section_rows = xl[xl.iloc[:, 1] == section].index
    xl.loc[section_rows, 'Section'] = section
    xl.loc[section_rows, xl.columns[1]] = ""  # Clear the original cells
    print(section)

# Remove all blank rows
xl = xl.dropna(how='all')




# Define the list of sections
sections = ['HARDWARE PARTS', 'ACCESSORY PARTS for BUYOUT', 'ADDITIONAL ACCESSORY PARTS',
            'RECESSED HARDWARE - Install Prior to Shipping', 'Berenson INTEGRATED PULL PARTS']

# Create a dictionary where the keys are section names and the values are lists of their respective row indexes
section_indexes = {}
for section in sections:
    section_indexes[section] = xl[xl[xl.columns[0]] == section].index.tolist()

# Iterate over the sections dictionary
for i, (section, indexes) in enumerate(section_indexes.items()):
    if not indexes:
        continue  # If the section doesn't exist, skip to the next iteration
    start_index = indexes[0]  # Start of the current section
    if i+1 < len(sections):  # If this isn't the last section
        next_section = sections[i+1]
        if next_section in section_indexes:  # If the next section exists
            end_index = section_indexes[next_section][0]  # Start of the next section
        else:
            end_index = xl.index[-1]  # If next section doesn't exist, set end_index to the end of the DataFrame
    else:
        end_index = xl.index[-1]  # If this is the last section, set end_index to the end of the DataFrame

    print(f"Processing section: {section}")  # Print the section name here

    if section == 'HARDWARE PARTS':
        # Apply logic for "HARDWARE PARTS" section here
        pass
    elif section == 'ACCESSORY PARTS for BUYOUT':
        xl.loc[start_index + 1 : end_index - 1, xl.columns[6]] = xl.loc[start_index + 1 : end_index - 1, xl.columns[1]]
        #xl.loc[start_index + 1 : end_index - 1, xl.columns[1]] = ""
        xl.loc[start_index + 1 : end_index - 1, xl.columns[1]] = xl.loc[start_index + 1 : end_index - 1, xl.columns[2]]
        xl.loc[start_index + 1 : end_index - 1, xl.columns[2]] = ""
        xl.loc[start_index + 1: end_index - 1, xl.columns[5]] = xl.loc[start_index + 1: end_index - 1, xl.columns[4]]
        xl.loc[start_index + 1: end_index - 1, xl.columns[4]] = ""
        print(section_rows)

    elif section == 'ADDITIONAL ACCESSORY PARTS':
    #
        # Apply logic for "ADDITIONAL ACCESSORY PARTS" section here
        # xl.loc[start_index + 1: end_index - 1, xl.columns[0]] = xl.loc[start_index + 1: end_index - 1, xl.columns[1]]

        xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1, xl.columns[3]]
        #xl.loc[start_index + 1: end_index - 1, xl.columns[3]] = ''
        #xl.loc[start_index + 1: end_index - 1, xl.columns[1]] = xl.loc[start_index + 1: end_index - 1, xl.columns[4]]
        #xl.loc[start_index + 1: end_index - 1, xl.columns[4]] = ''

    # elif section == 'Berenson INTEGRATED PULL PARTS':
    #     # Apply logic for "Berenson INTEGRATED PULL PARTS" section here
    #     xl.loc[start_index + 1: end_index - 1, xl.columns[2]] = xl.loc[start_index + 1: end_index - 1, xl.columns[0]]


    elif section == 'RECESSED HARDWARE - Install Prior to Shipping':
        # Apply logic for "RECESSED HARDWARE - Install Prior to Shipping" section here
        pass





# Remove all blank rows
xl = xl.dropna(how='all')

# Print the DataFrame to console
print(xl)

# Before saving the Excel file, call the final_cleanup function
# xl = final_cleanup(xl)

# Replace empty strings with NaN
xl.replace('', np.nan, inplace=True)

# Identify the rows where the first column contains 'PART NAME'
rows_to_drop_partname = xl[xl.iloc[:, 1] == 'PART NAME'].index

# Drop these rows
xl.drop(rows_to_drop_partname, inplace=True)

# Identify the rows where the first column contains 'PART NAME'
rows_to_drop_qty = xl[xl.iloc[:, 0] == 'QTY'].index
# Drop these rows
xl.drop(rows_to_drop_qty, inplace=True)

# Identify the rows where the first column contains 'PART NAME'
rows_to_drop_qty2 = xl[xl.iloc[:, 3] == 'QTY'].index
# Drop these rows
xl.drop(rows_to_drop_qty2, inplace=True)

# Identify the rows where the third column contains 'Description'
rows_to_drop_description = xl[xl.iloc[:, 2] == 'Description'].index

# Drop these rows
xl.drop(rows_to_drop_description, inplace=True)

# List of words to check
words = ['BUY', 'PICKED', 'ID#', 'PART #', 'PART  #', 'QTY', 'ID  #']

# Filter rows where any of the specified words appear in columns 7 and beyond
rows_to_remove = xl[xl.iloc[:, 7:].isin(words).any(axis=1)].index
#clear cells with words

# Remove these rows from the DataFrame
xl = xl.drop(rows_to_remove)

#wipe any remaining cells that contain words from list of words
xl = xl.replace(words, '')

# Then drop the rows that are entirely composed of NaN values
xl.dropna(how='all', inplace=True)

#remove all columns that are entirely blank
xl.dropna(axis=1, how='all', inplace=True)





# Reset the index
xl = xl.reset_index(drop=True)

# Save to new .xlsx file (in Excel 2007+ format)
xl.to_excel('cleaned_data.xlsx', index=False)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Transfer data from pandas DataFrame to the worksheet
for r in dataframe_to_rows(xl, index=False, header=True):
    ws.append(r)

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the workbook
wb.save('cleaned_data.xlsx')
