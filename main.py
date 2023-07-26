import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os


filename = 'am002456.xls'



def is_number(n):
    try:
        float(n)
        return True
    except ValueError:
        return False

#original_sheet = job summary-landscape.xls
# Load spreadsheet
xl = pd.read_excel(filename, engine='xlrd')

# Initial processing, getting rid of whole pages
start_index = xl[xl.apply(lambda row: row.astype(str).str.contains('HARDWARE PARTS').any(), axis=1)].index[0]
end_index = xl[xl.apply(lambda row: row.astype(str).str.contains('Packsize Program').any(), axis=1)].index[0]
metal_parts_index = xl[xl.apply(lambda row: row.astype(str).str.contains('Metal Parts - Cut Length and Qty').any(), axis=1)].index[0]
accessory_buyout_index = xl[xl.apply(lambda row: row.astype(str).str.contains('ACCESSORY PARTS for BUYOUT').any(), axis=1)].index[0]

xl = xl.loc[start_index:end_index-1]
xl = xl.drop(list(range(metal_parts_index, accessory_buyout_index)))

# Find rows where the 4th column is either NaN, empty, or contains only whitespace
# For these rows, replace the value in the 4th column with the value from the 3rd column
xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()), xl.columns[3]] = xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()), xl.columns[2]]

# Then, for these same rows, set the value in the 3rd column to an empty string
xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()), xl.columns[2]] = ""


# Section handling
sections = [
    'HARDWARE PARTS',
    'Hinges & Mounting Plates',
    'Legrabox & Antaro',
    'Metabox, Tandem & Accuride',
    'Blum Metal Parts',
    'Closet',
    'Other',
    'ACCESSORY PARTS for BUYOUT',
    'ADDITIONAL ACCESSORY PARTS',
    'Berenson INTEGRATED PULL PARTS',
    'RECESSED HARDWARE - Install Prior to Shipping',
]

xl.insert(0, 'Section', np.nan)

for section in sections:
    section_rows = xl[xl.iloc[:, 1] == section].index
    xl.loc[section_rows, 'Section'] = section
    xl.loc[section_rows, xl.columns[1]] = ""
    print(section)

section_indexes = {}
for section in sections:
    section_indexes[section] = xl[xl[xl.columns[0]] == section].index.tolist()

for i, (section, indexes) in enumerate(section_indexes.items()):
    if not indexes:
        continue
    start_index = indexes[0]
    if i+1 < len(sections):
        next_section = sections[i+1]
        if next_section in section_indexes and section_indexes[next_section]: # check if the section exists and is not empty
            end_index = section_indexes[next_section][0]
        else:
            end_index = xl.index[-1]
    else:
        end_index = xl.index[-1]
    print(f"Start index for {section}: {start_index}")
    print(f"End index for {section}: {end_index}")
    print(f"Processing section: {section}")

    if section == 'HARDWARE PARTS':
        pass
    elif section == 'ACCESSORY PARTS for BUYOUT':
        xl.loc[start_index + 1 : end_index - 1, xl.columns[4]] = xl.loc[start_index + 1 : end_index - 1, xl.columns[5]]
        # xl.loc[start_index + 1 : end_index - 1, xl.columns[2]] = ""
        xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1, xl.columns[1]]
        xl.loc[start_index + 1 : end_index - 1, xl.columns[1]] = xl.loc[start_index + 1 : end_index - 1, xl.columns[2]]
        # xl.loc[start_index + 1: end_index - 1, xl.columns[4]] = ""

    elif section == 'Blum Metal Parts':
        xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1, xl.columns[7]]
        xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()),
        xl.columns[3]] = xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()), xl.columns[2]]
        xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()), xl.columns[2]] = ""
    elif section == 'Closet':
        xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1, xl.columns[7]]

    elif section == 'Other':
        # create a mask for NaN or empty or blank cells in column 6
        mask = xl[xl.columns[6]].isna() | (xl[xl.columns[6]] == '') | (xl[xl.columns[6]].astype(str).str.isspace())
        # apply the mask along with the original condition
        xl.loc[mask & (start_index + 1 <= xl.index) & (xl.index <= end_index - 1), xl.columns[6]] = xl.loc[
        mask & (start_index + 1 <= xl.index) & (xl.index <= end_index - 1), xl.columns[7]]

    elif section == 'ADDITIONAL ACCESSORY PARTS':
             xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1, xl.columns[3]]
             xl.loc[start_index + 1: end_index - 1, xl.columns[1]] = xl.loc[start_index + 1: end_index - 1, xl.columns[4]]

    elif section == 'RECESSED HARDWARE - Install Prior to Shipping':
            pass

    elif section == 'Berenson INTEGRATED PULL PARTS':
        xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1, xl.columns[7]]



xl = xl.dropna(how='all')



print(xl)
# final touchups to the dataframe, replacing NaNs with empty strings, removing rows based on conditions, etc.
xl.replace('', np.nan, inplace=True)
rows_to_drop_partname = xl[xl.iloc[:, 1] == 'PART NAME'].index
xl.drop(rows_to_drop_partname, inplace=True)
rows_to_drop_qty = xl[xl.iloc[:, 0] == 'QTY'].index
xl.drop(rows_to_drop_qty, inplace=True)
rows_to_drop_qty2 = xl[xl.iloc[:, 3] == 'QTY'].index
xl.drop(rows_to_drop_qty2, inplace=True)
rows_to_drop_description = xl[xl.iloc[:, 2] == 'Description'].index
xl.drop(rows_to_drop_description, inplace=True)
words = ['BUY', 'PICKED', 'ID#', 'PART #', 'PART  #', 'QTY', 'ID  #']
rows_to_remove = xl[xl.iloc[:, 7:].isin(words).any(axis=1)].index
xl = xl.drop(rows_to_remove)
xl = xl.replace(words, '')
print(xl.columns)

# Replace 'Custom Drilling- See CDR form' with an empty string '' in column B
xl['Unnamed: 0'] = xl['Unnamed: 0'].str.replace('\s*Custom Drilling- See CDR form', '', regex=True)

xl.dropna(how='all', inplace=True)

xl = xl.reset_index(drop=True)
#xl.drop(xl.columns[[2, 4, 6, 7]], axis=1, inplace=True)

xl.to_excel('cleaned_data3.xlsx', index=False)

# Assuming filename contains the name of the original file
file_name, _ = os.path.splitext(filename)

# Append "_cleaned" and add the file extension ".xlsx"
new_file_name = f"{file_name}_cleaned.xlsx"

# Now you can save the DataFrame to the new Excel file
xl.to_excel(new_file_name, index=False)

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Transfer data from pandas DataFrame to the worksheet
for r in dataframe_to_rows(xl, index=False, header=True):
    ws.append(r)
# Rename columns
ws['A1'] = 'SECTION'
ws['B1'] = 'Inventory ID'
ws['E1'] = 'DESCRIPTION'
ws['G1'] = 'QTY'
from openpyxl.utils import column_index_from_string

# The columns to keep
cols_to_keep = ['A', 'B', 'E', 'G']

# Convert column letters to indices
indices_to_keep = [column_index_from_string(col) for col in cols_to_keep]

# Delete columns in reverse order, so that deleting one column doesn't shift the indices of the others
for col_idx in range(ws.max_column, 0, -1):  # start from the last column
    if col_idx not in indices_to_keep:
        ws.delete_cols(col_idx)


# Adjust column widths
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the workbook
wb.save(new_file_name)
