
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import math

def clean_excel_file(filename):
    # Load spreadsheet
    xl = pd.read_excel(filename, engine='xlrd')
    def is_number(n):
        try:
            float(n)
            return True
        except ValueError:
            return False

    # original_sheet = job summary-landscape.xls
    # Load spreadsheet
    xl = pd.read_excel(filename, engine='xlrd')

    # Initial processing, getting rid of whole pages
    start_index = xl[xl.apply(lambda row: row.astype(str).str.contains('HARDWARE PARTS').any(), axis=1)].index[0]
    end_index = xl[xl.apply(lambda row: row.astype(str).str.contains('Packsize Program').any(), axis=1)].index[0]
    metal_parts_index = \
    xl[xl.apply(lambda row: row.astype(str).str.contains('Metal Parts - Cut Length and Qty').any(), axis=1)].index[0]
    accessory_buyout_index = \
    xl[xl.apply(lambda row: row.astype(str).str.contains('ACCESSORY PARTS for BUYOUT').any(), axis=1)].index[0]

    xl = xl.loc[start_index:end_index - 1]
    xl = xl.drop(list(range(metal_parts_index, accessory_buyout_index)))

    # Find rows where the 4th column is either NaN, empty, or contains only whitespace
    # For these rows, replace the value in the 4th column with the value from the 3rd column
    xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()),
    xl.columns[3]] = xl.loc[
        xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()),
        xl.columns[2]]

    # Then, for these same rows, set the value in the 3rd column to an empty string
    xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()),
    xl.columns[2]] = ""

    # Section handling
    sections = [
        'HARDWARE PARTS',
        'Hinges & Mounting Plates',
        'Legrabox & Antaro',
        'Metabox, Tandem & Accuride',
        'Blum Metal Parts',
        'Closet',
        'Other',
        'ACCESSORY PARTS for BUYOUT',
        'ADDITIONAL ACCESSORY PARTS',
        'Berenson INTEGRATED PULL PARTS',
        'RECESSED HARDWARE - Install Prior to Shipping',
    ]

    xl.insert(0, 'Section', np.nan)

    for section in sections:
        section_rows = xl[xl.iloc[:, 1] == section].index
        xl.loc[section_rows, 'Section'] = section
        xl.loc[section_rows, xl.columns[1]] = ""
        print(section)

    section_indexes = {}
    for section in sections:
        section_indexes[section] = xl[xl[xl.columns[0]] == section].index.tolist()

    for i, (section, indexes) in enumerate(section_indexes.items()):
        if not indexes:
            continue
        start_index = indexes[0]
        if i + 1 < len(sections):
            next_section = sections[i + 1]
            if next_section in section_indexes and section_indexes[
                next_section]:  # check if the section exists and is not empty
                end_index = section_indexes[next_section][0]
            else:
                end_index = xl.index[-1]
        else:
            end_index = xl.index[-1]
        print(f"Start index for {section}: {start_index}")
        print(f"End index for {section}: {end_index}")
        print(f"Processing section: {section}")

        if section == 'HARDWARE PARTS':
            pass
        elif section == 'ACCESSORY PARTS for BUYOUT':
            xl.loc[start_index + 1: end_index - 1, xl.columns[4]] = xl.loc[start_index + 1: end_index - 1,
                                                                    xl.columns[5]]
            # xl.loc[start_index + 1 : end_index - 1, xl.columns[2]] = ""
            xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1,
                                                                    xl.columns[1]]
            xl.loc[start_index + 1: end_index - 1, xl.columns[1]] = xl.loc[start_index + 1: end_index - 1,
                                                                    xl.columns[2]]
            # xl.loc[start_index + 1: end_index - 1, xl.columns[4]] = ""

        elif section == 'Blum Metal Parts':
            xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1,
                                                                    xl.columns[7]]
            xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()),
            xl.columns[3]] = xl.loc[
                xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()),
                xl.columns[2]]
            xl.loc[xl[xl.columns[3]].isna() | (xl[xl.columns[3]] == '') | (xl[xl.columns[3]].astype(str).str.isspace()),
            xl.columns[2]] = ""
        elif section == 'Closet':
            xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1,
                                                                    xl.columns[7]]

        elif section == 'Other':
            # create a mask for NaN or empty or blank cells in column 6
            mask = xl[xl.columns[6]].isna() | (xl[xl.columns[6]] == '') | (xl[xl.columns[6]].astype(str).str.isspace())
            # apply the mask along with the original condition
            xl.loc[mask & (start_index + 1 <= xl.index) & (xl.index <= end_index - 1), xl.columns[6]] = xl.loc[
                mask & (start_index + 1 <= xl.index) & (xl.index <= end_index - 1), xl.columns[7]]

        elif section == 'ADDITIONAL ACCESSORY PARTS':
            xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1,
                                                                    xl.columns[3]]
            xl.loc[start_index + 1: end_index - 1, xl.columns[1]] = xl.loc[start_index + 1: end_index - 1,
                                                                    xl.columns[4]]

        elif section == 'RECESSED HARDWARE - Install Prior to Shipping':
            pass

        elif section == 'Berenson INTEGRATED PULL PARTS':
            xl.loc[start_index + 1: end_index - 1, xl.columns[6]] = xl.loc[start_index + 1: end_index - 1,
                                                                    xl.columns[7]]

    xl = xl.dropna(how='all')

    print(xl)
    # final touchups to the dataframe, replacing NaNs with empty strings, removing rows based on conditions, etc.
    xl.replace('', np.nan, inplace=True)
    rows_to_drop_partname = xl[xl.iloc[:, 1] == 'PART NAME'].index
    xl.drop(rows_to_drop_partname, inplace=True)
    rows_to_drop_qty = xl[xl.iloc[:, 0] == 'QTY'].index
    xl.drop(rows_to_drop_qty, inplace=True)
    rows_to_drop_qty2 = xl[xl.iloc[:, 3] == 'QTY'].index
    xl.drop(rows_to_drop_qty2, inplace=True)
    rows_to_drop_description = xl[xl.iloc[:, 2] == 'Description'].index
    xl.drop(rows_to_drop_description, inplace=True)
    words = ['BUY', 'PICKED', 'ID#', 'PART #', 'PART  #', 'QTY', 'ID  #']
    rows_to_remove = xl[xl.iloc[:, 7:].isin(words).any(axis=1)].index
    xl = xl.drop(rows_to_remove)
    xl = xl.replace(words, '')
    print(xl.columns)

    # Replace 'Custom Drilling- See CDR form' with an empty string '' in column B
    xl['Unnamed: 0'] = xl['Unnamed: 0'].str.replace('\s*Custom Drilling- See CDR form', '', regex=True)

    xl.dropna(how='all', inplace=True)

    xl = xl.reset_index(drop=True)
    # xl.drop(xl.columns[[2, 4, 6, 7]], axis=1, inplace=True)

    xl.to_excel('cleaned_data3.xlsx', index=False)

    # Assuming filename contains the name of the original file
    file_name, _ = os.path.splitext(filename)

    # Append "_cleaned" and add the file extension ".xlsx"
    new_file_name = f"{file_name}_cleaned.xlsx"

    # Now you can save the DataFrame to the new Excel file
    # xl.to_excel(new_file_name, index=False)

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Remove rows where 'Inventory ID' is empty
    xl = xl[pd.notnull(xl.iloc[:, 1])]  # Selects second column for check (0-indexed)
    xl.reset_index(drop=True, inplace=True)

    # Transfer data from pandas DataFrame to the worksheet
    for r in dataframe_to_rows(xl, index=False, header=True):
        ws.append(r)
    # Rename columns

    ws['A1'].value = 'SECTION'
    ws['B1'].value = 'Inventory ID'
    ws['E1'].value = 'Description'
    ws['G1'].value = 'Qty Required'

    from openpyxl.utils import column_index_from_string, get_column_letter

    # The columns to keep
    cols_to_keep = ['A', 'B', 'E', 'G']

    # Convert column letters to indices
    indices_to_keep = [column_index_from_string(col) - 1 for col in cols_to_keep]

    # Delete columns in reverse order, so that deleting one column doesn't shift the indices of the others
    for col_idx in range(ws.max_column, -1, -1):  # start from the last column
        if col_idx not in indices_to_keep:
            ws.delete_cols(col_idx+1)

    # Set width for column A to 30
    ws.column_dimensions['A'].width = 30

    # Set width for column B to 20
    ws.column_dimensions['B'].width = 40

    # Set width for column C to 40
    ws.column_dimensions['C'].width = 13

    # remove 'Section' column
    ws.delete_cols(1)



    # Assume ws is your worksheet object
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        cell = row[2]  # change the index based on your 'Qty Required' column index
        if cell.value is not None and isinstance(cell.value, float):
            cell.value = math.ceil(cell.value)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace('_', ' ')

    # Load the template
    wb_template = load_workbook('BOM Template.xlsx')
    ws_template = wb_template.active

    # Convert cleaned data to a DataFrame
    df_clean = pd.DataFrame(ws.values)

    # Drop the header row from cleaned data
    df_clean = df_clean.iloc[1:]

    # Get the number of non-empty rows in the template before appending
    rows_in_template = sum((1 for row in ws_template.iter_rows() if any(cell.value for cell in row)))

    # Append data from the cleaned DataFrame to the template worksheet
    for index, row in df_clean.iterrows():
        for i, value in enumerate(row.values):
            # index + 2 to offset zero-based index and header row
            # i + 1 to offset zero-based index
            ws_template.cell(row=index + rows_in_template + 1, column=i + 1, value=value)

    # Save the template
    wb_template.save(new_file_name)

    return xl

def process_files_in_directory(dir_path):
    # Get list of .xls files in the directory
    files = [f for f in os.listdir(dir_path) if f.endswith('.xls')]

    # Process all .xls files
    for file in files:
        # Generate file path
        file_path = os.path.join(dir_path, file)

        # Process file
        clean_excel_file(file_path)

        # Save to a new Excel file
        new_file_name = f"{os.path.splitext(file)[0]}_cleaned.xlsx"
        new_file_path = os.path.join(dir_path, new_file_name)
        # df.to_excel(new_file_path, index=False)

if __name__ == "__main__":
    # Print a statement to ensure this block is being reached
    print("Preparing to process files.")
    # Get user input for directory
    dir_path = input("Please enter the path of the directory: ")
    process_files_in_directory(dir_path)
